"""Report builder regression tests (6-sheet Korean Excel output).

Validates scripts.report_builder.build_report:
- six Korean sheets are emitted
- header row fill colour (4472C4)
- Reject finding -> FAIL verdict cell fill (FFC7CE)
- C2/C8 provenance: evidence source_file appears in 지적사항 종합 출처 column
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from scripts.report_builder import build_report

_CASE_ID = "4"
_CERT_PDF = "PU2405873-W2411008-SEONGHWA-1-21PCS MTC_cert.pdf"


def _minimal_manifest() -> dict:
    """Self-contained manifest (case '4') — no dependency on a generated
    manifest.json or a fixed checkout path."""
    return {
        "schema_version": "2.0",
        "case_count": 1,
        "cases": [
            {
                "case_id": _CASE_ID,
                "cert_dir": f"standard inspection Cert cleanup data/{_CASE_ID}",
                "mps_dir": None,
                "cert_pdfs": [_CERT_PDF],
                "mps_pdfs": [],
                "has_cert_pdf": True,
                "has_mps": False,
                "is_zip_only": False,
            }
        ],
    }

# A distinctive substring used to assert provenance round-trips into the sheet.
_CHEM_SOURCE_FILE = "PU2405873-W2411008-SEONGHWA-1-21PCS MTC_cert.pdf"

_EXPECTED_SHEETS = {
    "검토 총괄",
    "화학성분 검토",
    "기계적성질 검토",
    "열처리 검토",
    "표기·형식 검토",
    "지적사항 종합",
}


def _evidence(channel: str, source_file: str, anchor: str, snippet: str) -> dict:
    """Build a 4-field provenance evidence entry (C2/C8)."""
    return {
        "channel": channel,
        "source_file": source_file,
        "anchor": anchor,
        "snippet": snippet,
        "sha256": "a" * 64,
    }


def _synthetic_findings() -> dict:
    """A compare_engine-shaped dict covering 5 categories incl. a Reject."""
    return {
        "case_id": _CASE_ID,
        "findings": [
            {
                "finding_id": "CHEM-MPS-001",
                "category": "Chemistry",
                "severity": "Reject",
                "material_grade": "P91",
                "heat_no": "H12345",
                "cert_pdf": _CERT_PDF,
                "page_ref": "p.3",
                "issue_summary": "Cr below MPS minimum",
                "details": (
                    "Cr measured 7.80% is below the MPS restricted minimum "
                    "of 8.00% for SA-335 P91."
                ),
                "structured": {
                    "element": "Cr",
                    "measured": 7.80,
                    "min": 8.00,
                    "max": 9.50,
                    "unit": "%",
                    "property": None,
                    "stage": None,
                    "spec": "SA-335 P91",
                    "location": None,
                    "mtc_value": None,
                    "correct_value": None,
                    "item_name": "Seamless Pipe P91",
                },
                "required_action": "Reject heat; request re-test or replacement.",
                "evidence": [
                    _evidence(
                        "body",
                        _CHEM_SOURCE_FILE,
                        "p.3#chem-Cr",
                        "Cr 7.80",
                    ),
                    _evidence(
                        "mps",
                        "16. MPS-ANDRES-335P91.pdf",
                        "p.2#chem",
                        "Cr 8.00-9.50",
                    ),
                ],
            },
            {
                "finding_id": "MECH-001",
                "category": "Mechanical",
                "severity": "ActionRequired",
                "material_grade": "P91",
                "heat_no": "H12345",
                "cert_pdf": _CERT_PDF,
                "page_ref": "p.4",
                "issue_summary": "Tensile strength below minimum",
                "details": "Tensile strength 580 MPa is below the 585 MPa minimum.",
                "structured": {
                    "element": None,
                    "measured": 580,
                    "min": 585,
                    "max": None,
                    "unit": "MPa",
                    "property": "Tensile Strength",
                    "stage": None,
                    "spec": "min 585 MPa",
                    "location": None,
                    "mtc_value": None,
                    "correct_value": None,
                    "item_name": "Seamless Pipe P91",
                },
                "required_action": "Request clarification / re-test.",
                "evidence": [
                    _evidence(
                        "body",
                        _CERT_PDF,
                        "p.4#tensile",
                        "TS 580",
                    ),
                ],
            },
            {
                "finding_id": "HT-001",
                "category": "HeatTreatment",
                "severity": "Question",
                "material_grade": "P91",
                "heat_no": "H12345",
                "cert_pdf": _CERT_PDF,
                "page_ref": "p.5",
                "issue_summary": "Tempering temperature not stated",
                "details": "PWHT tempering temperature is missing from the MTC.",
                "structured": {
                    "element": None,
                    "measured": None,
                    "min": 730,
                    "max": 800,
                    "unit": "C",
                    "property": None,
                    "stage": "Tempering",
                    "spec": "730-800 C",
                    "location": None,
                    "mtc_value": None,
                    "correct_value": None,
                    "item_name": "Seamless Pipe P91",
                },
                "required_action": "Confirm tempering temperature with mill.",
                "evidence": [
                    _evidence(
                        "ref_code",
                        "SA-335.md",
                        "p.10#pwht",
                        "730 to 800",
                    ),
                ],
            },
            {
                "finding_id": "ID-001",
                "category": "Identification",
                "severity": "Minor",
                "material_grade": "P91",
                "heat_no": "H12345",
                "cert_pdf": _CERT_PDF,
                "page_ref": "p.1",
                "issue_summary": "Heat number formatting note",
                "details": "Heat number printed without leading zero.",
                "structured": {
                    "element": None,
                    "measured": None,
                    "min": None,
                    "max": None,
                    "unit": None,
                    "property": None,
                    "stage": None,
                    "spec": None,
                    "location": "Heat No field",
                    "mtc_value": "12345",
                    "correct_value": "H12345",
                    "item_name": "Seamless Pipe P91",
                },
                "required_action": "Verify heat number against mill record.",
                "evidence": [
                    _evidence(
                        "body",
                        _CERT_PDF,
                        "p.1#heatno",
                        "12345",
                    ),
                ],
            },
            {
                "finding_id": "DOC-001",
                "category": "DocumentError",
                "severity": "ActionRequired",
                "material_grade": "P91",
                "heat_no": "H12345",
                "cert_pdf": _CERT_PDF,
                "page_ref": "p.1",
                "issue_summary": "Cert number typo",
                "details": "Certificate number on page 1 differs from page 2.",
                "structured": {
                    "element": None,
                    "measured": None,
                    "min": None,
                    "max": None,
                    "unit": None,
                    "property": None,
                    "stage": None,
                    "spec": None,
                    "location": "Certificate No field",
                    "mtc_value": "C-2024-001",
                    "correct_value": "C-2024-0001",
                    "item_name": "Seamless Pipe P91",
                },
                "required_action": "Request corrected reissue from mill.",
                "evidence": [
                    _evidence(
                        "body",
                        _CERT_PDF,
                        "p.1#certno",
                        "C-2024-001",
                    ),
                ],
            },
        ],
    }


@pytest.fixture()
def report_path(tmp_path: Path) -> Path:
    """Write synthetic findings, build the report, return the xlsx path."""
    findings_path = tmp_path / f"{_CASE_ID}_findings.json"
    findings_path.write_text(
        json.dumps(_synthetic_findings(), ensure_ascii=False),
        encoding="utf-8",
    )
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(
        json.dumps(_minimal_manifest(), ensure_ascii=False), encoding="utf-8"
    )
    out_dir = tmp_path / "output"
    out_path = build_report(
        case_id=_CASE_ID,
        findings_path=findings_path,
        manifest_path=manifest_path,
        out_dir=out_dir,
    )
    assert out_path.exists(), "build_report did not produce an output file"
    return out_path


def _fg(cell) -> str:
    """Return the cell fill foreground colour as an upper-case rgb string."""
    fill = cell.fill
    if fill is None or fill.fgColor is None:
        return ""
    rgb = fill.fgColor.rgb
    if not isinstance(rgb, str):
        return ""
    return rgb.upper()


def test_six_sheets_exist(report_path: Path):
    wb = openpyxl.load_workbook(report_path)
    assert set(wb.sheetnames) == _EXPECTED_SHEETS


def test_header_fill_color(report_path: Path):
    wb = openpyxl.load_workbook(report_path)
    ws = wb["검토 총괄"]
    fg = _fg(ws["A4"])  # header row is row 4 on the summary sheet
    assert fg.endswith("4472C4"), (
        f"summary header A4 fill should end with 4472C4, got {fg!r}"
    )


def test_fail_cell_color(report_path: Path):
    """A Reject finding must yield a FAIL verdict cell filled FFC7CE."""
    wb = openpyxl.load_workbook(report_path)

    found_fail_fill = False
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if "FFC7CE" in _fg(cell):
                    found_fail_fill = True
                    break
            if found_fail_fill:
                break
        if found_fail_fill:
            break

    assert found_fail_fill, (
        "expected at least one FAIL cell with fill fgColor containing FFC7CE"
    )


def test_provenance_in_findings_sheet(report_path: Path):
    """C2/C8: 출처 (last) column carries the evidence source_file substring."""
    wb = openpyxl.load_workbook(report_path)
    ws = wb["지적사항 종합"]

    # The 출처 header sits in the last populated header column (row 3).
    header_row = 3
    last_col = ws.max_column
    assert ws.cell(row=header_row, column=last_col).value == "출처", (
        "last header column on 지적사항 종합 should be 출처, got "
        f"{ws.cell(row=header_row, column=last_col).value!r}"
    )

    # Collect every 출처 value in the data rows.
    source_values = [
        str(ws.cell(row=r, column=last_col).value or "")
        for r in range(header_row + 1, ws.max_row + 1)
    ]
    joined = "\n".join(source_values)

    assert _CHEM_SOURCE_FILE in joined, (
        "expected evidence source_file substring "
        f"{_CHEM_SOURCE_FILE!r} in 출처 column values, got: {source_values!r}"
    )
