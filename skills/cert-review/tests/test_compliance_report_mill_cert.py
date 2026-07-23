"""T21/T22: compliance_report renders the conditional mill-cert sheet (기준 21·22).

Self-contained: builds a synthetic review.json under tmp_path, renders the xlsx,
and reads it back with openpyxl (Korean read-back). No dataset required.
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from scripts.compliance_report import build_compliance_report

_SHEET = "원자재 MILL CERT 검토"


def _base_review() -> dict:
    return {
        "case_id": "PU",
        "po_number": "PU2601564",
        "mps_files": ["mps_a.pdf"],
        "code_edition_note": "",
        "materials": [
            {
                "item_name": "Forged Elbow (PO Item No. 001)",
                "heat_no": "B14339",
                "grade_cert": "A182-F22 CL.3",
                "grade_spec": "SA-182-F22",
                "size": "DN15 SW",
                "qty": "33",
                "verdict": "FAIL",
                "chemistry": [], "mechanical": [], "heat_treatment": [],
                "nde": [], "doc_checks": [],
            }
        ],
        "findings": [],
    }


def _mill_cert_rows() -> list[dict]:
    return [
        {"item": "연결성: Heat No.", "source": "MILL CERT p.4 / MTC p.3",
         "mill_value": "B14339", "mtc_value": "B14339", "verdict": "PASS",
         "note": "정확 일치"},
        {"item": "인장 TS(MPa) 교차비교", "source": "MILL CERT p.4 / MTC p.3",
         "mill_value": "582.71", "mtc_value": "582.71", "verdict": "FAIL",
         "note": "소수점 둘째 자리까지 동일 — 완제품 시험 미실시(전사 복제) 의심"},
    ]


def _flat_cells(xlsx: Path, sheet: str) -> list[str]:
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[sheet]
    return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]


def test_mill_cert_sheet_rendered_with_fail_fill(tmp_path: Path):     # T21 / AC-7
    review = _base_review()
    review["materials"][0]["mill_cert"] = _mill_cert_rows()
    review_path = tmp_path / "PU_review.json"
    review_path.write_text(json.dumps(review, ensure_ascii=False), encoding="utf-8")

    out = build_compliance_report(review_path, tmp_path / "out.xlsx")

    wb = openpyxl.load_workbook(out)
    assert _SHEET in wb.sheetnames
    ws = wb[_SHEET]

    flat = _flat_cells(out, _SHEET)
    assert any("원자재 성적서(MILL CERT) 검증·교차비교 (기준 21·22)" in c for c in flat)
    assert any("인장 TS(MPa) 교차비교" in c for c in flat)
    assert any("연결성: Heat No." in c for c in flat)
    assert any("전사 복제" in c for c in flat)
    # 한국어 read-back 무결 (U+FFFD·`占` 부재).
    assert not any("�" in c or "占" in c for c in flat)

    # FAIL 셀 색 (verdict col 6): "FFFFC7CE".
    fail_fills = [
        ws.cell(row=r, column=6).fill.start_color.rgb
        for r in range(4, ws.max_row + 1)
        if ws.cell(row=r, column=6).value == "FAIL"
    ]
    assert fail_fills and all(f == "FFFFC7CE" for f in fail_fills)


def test_legacy_review_without_mill_cert_key_no_sheet(tmp_path: Path):   # T22 / AC-7
    review = _base_review()  # materials[0]에 mill_cert 키 없음 (legacy)
    review_path = tmp_path / "PU_review.json"
    review_path.write_text(json.dumps(review, ensure_ascii=False), encoding="utf-8")

    out = build_compliance_report(review_path, tmp_path / "out.xlsx")

    wb = openpyxl.load_workbook(out)
    assert _SHEET not in wb.sheetnames
    # 기존 6시트 렌더 불변.
    assert wb.sheetnames == [
        "검토 총괄", "화학성분 검토", "기계적성질 검토", "열처리 검토",
        "표기·형식 검토", "지적사항 종합",
    ]
    flat = _flat_cells(out, "검토 총괄")
    assert any("Forged Elbow" in c for c in flat)


def test_empty_mill_cert_array_no_sheet(tmp_path: Path):
    """mill_cert가 있어도 전부 빈 배열이면 시트 미생성 (조건부)."""
    review = _base_review()
    review["materials"][0]["mill_cert"] = []
    review_path = tmp_path / "PU_review.json"
    review_path.write_text(json.dumps(review, ensure_ascii=False), encoding="utf-8")

    out = build_compliance_report(review_path, tmp_path / "out.xlsx")
    wb = openpyxl.load_workbook(out)
    assert _SHEET not in wb.sheetnames
