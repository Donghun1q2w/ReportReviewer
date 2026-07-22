"""Deterministic compliance-review report (6-sheet Korean Excel).

Input: a structured review JSON (Cert vs MPS vs ref_code/Code comparison) — see
schema below. Unlike report_builder (which renders only findings/violations),
this renders the FULL per-parameter comparison (PASS rows included), which is
what a spec-compliance MTC review report shows.

NO OCR libraries are imported (C1) — openpyxl only.

Review JSON schema
------------------
{
  "case_id": "24",
  "po_number": "PU2501275",                 # optional
  "mps_files": ["1.MPS-SH-ACOR-106_0.pdf"],  # optional
  "code_edition_note": "성적서 2019 vs MPS 2022/2023",  # optional
  "materials": [
    {
      "item_name": "Seamless Pipe",
      "heat_no": "622090314",
      "grade_cert": "A106 Gr.B",            # as printed on the MTC
      "grade_spec": "SA-106-B",             # ordered/spec grade
      "size": "48.3 x 7.14",                # optional
      "qty": 120,                            # optional
      "verdict": "FAIL",                    # PASS | FAIL | 주의 | N/A (report canonicalises any variant)
      "chemistry":      [ {"element","analysis"?, "cert","spec_range","source","verdict","note"?} ],
      "mechanical":     [ {"property","cert","spec","source","verdict","note"?} ],
      "heat_treatment": [ {"stage","cert","spec","source","verdict","note"?} ],
      "nde":            [ {"item","spec"?, "cert"?, "source"?, "verdict","note"?} ],
      "doc_checks":     [ {"page"?, "location","mtc_value","expected","verdict","note"?} ]
    }
  ],
  "findings": [ {"no","severity","category","location","content","action"} ]
}
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="FF4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_TITLE_FONT = Font(bold=True)
_PASS_FILL = PatternFill("solid", fgColor="FFC6EFCE")
_FAIL_FILL = PatternFill("solid", fgColor="FFFFC7CE")
_WARN_FILL = PatternFill("solid", fgColor="FFFFEB9C")
_NA_FILL = PatternFill("solid", fgColor="FFD9D9D9")

# --- Verdict / severity canonicalisation -------------------------------------
# Sub-agents emit free verdict/severity vocabulary ("합격", "ActionRequired",
# "정보성", "REVIEW", "N/A", …). To keep the report self-consistent, every cell
# verdict is collapsed to a fixed display set {PASS | FAIL | 주의 | N/A} and a
# colour is ALWAYS applied (no blank verdict cells). The keys below are matched
# case-insensitively (looked up verbatim first, then upper-cased).

# Cell verdict canonical set: PASS | FAIL | 주의 | N/A
_VERDICT_ALIASES = {
    # PASS
    "PASS": "PASS", "합격": "PASS", "적합": "PASS", "OK": "PASS",
    # FAIL
    "FAIL": "FAIL", "불합격": "FAIL", "부적합": "FAIL", "REJECT": "FAIL",
    "DOCUMENTERROR": "FAIL", "DOCUMENT ERROR": "FAIL", "기준 미달": "FAIL",
    "기준값 초과": "FAIL",
    # 주의 (warning / review / conditional / action-required leaking in as verdict)
    "주의": "주의", "WARNING": "주의", "WARN": "주의", "REVIEW": "주의",
    "QUESTION": "주의", "ACTIONREQUIRED": "주의", "ACTION REQUIRED": "주의",
    "MINOR": "주의", "PASS_WITH_FINDINGS": "주의",
    "조건부 PASS": "주의", "조건부PASS": "주의", "조건부합격": "주의",
    # N/A (not-applicable / informational)
    "N/A": "N/A", "NA": "N/A", "N.A.": "N/A", "INFO": "N/A", "INFORMATION": "N/A",
    "정보": "N/A", "정보성": "N/A", "참고": "N/A",
    "확인 불가": "N/A", "확인불가": "N/A", "미판정": "N/A",
}
_VERDICT_FILL = {
    "PASS": _PASS_FILL, "FAIL": _FAIL_FILL, "주의": _WARN_FILL, "N/A": _NA_FILL,
}

# Finding severity canonical set: Reject | ActionRequired | Question | Minor | Info
_SEVERITY_ALIASES = {
    "REJECT": "Reject", "FAIL": "Reject", "불합격": "Reject", "부적합": "Reject",
    "ACTIONREQUIRED": "ActionRequired", "ACTION REQUIRED": "ActionRequired",
    "DOCUMENTERROR": "ActionRequired", "DOCUMENT ERROR": "ActionRequired",
    "QUESTION": "Question", "주의": "Question", "WARNING": "Question",
    "MINOR": "Minor",
    "INFO": "Info", "INFORMATION": "Info",
    "정보": "Info", "정보성": "Info", "참고": "Info",
}
_SEVERITY_FILL = {
    "Reject": _FAIL_FILL, "ActionRequired": _FAIL_FILL,
    "Question": _WARN_FILL, "Minor": _WARN_FILL, "Info": _NA_FILL,
}


def _canon_verdict(v: Any) -> str:
    """Collapse any verdict variant to one of PASS | FAIL | 주의 | N/A.

    Empty/None → N/A. Unrecognised non-empty token → 주의 (surfaced for review
    rather than silently shown as a benign N/A)."""
    if v is None:
        return "N/A"
    key = str(v).strip()
    if not key:
        return "N/A"
    return _VERDICT_ALIASES.get(key) or _VERDICT_ALIASES.get(key.upper()) or "주의"


def _canon_severity(s: Any) -> str:
    """Collapse a finding severity to the canonical 5-level set; unknown
    non-empty tokens are kept verbatim (and shaded grey)."""
    if s is None:
        return ""
    key = str(s).strip()
    if not key:
        return ""
    return _SEVERITY_ALIASES.get(key) or _SEVERITY_ALIASES.get(key.upper()) or key


def _verdict_fill(v: Any) -> PatternFill:
    return _VERDICT_FILL.get(_canon_verdict(v), _NA_FILL)


def _fill_for(verdict: Any) -> PatternFill:
    """Verdict-cell fill (canonical PASS/FAIL/주의/N/A); always a solid colour."""
    return _VERDICT_FILL.get(_canon_verdict(verdict), _NA_FILL)


def _severity_fill(s: Any) -> PatternFill:
    return _SEVERITY_FILL.get(_canon_severity(s), _NA_FILL)


def _title(ws, text: str, ncols: int) -> None:
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = text
    c.font = _TITLE_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20


def _meta(ws, row: int, text: str, ncols: int) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    ws.cell(row=row, column=1).value = text


def _header(ws, row: int, headers: list[str], widths: list[float]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i)
        c.value = h
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22


def _mat_label(m: dict) -> str:
    name = m.get("item_name") or "—"
    heat = m.get("heat_no") or ""
    grade = m.get("grade_spec") or m.get("grade_cert") or ""
    tail = ", ".join(x for x in (heat, grade) if x)
    return f"{name}\n({tail})" if tail else name


def _grouped_sheet(wb, sheet, title, headers, widths, materials, key,
                   row_fields) -> None:
    """Build a sheet whose rows come from materials[*][key]; col A merges per material."""
    ws = wb.create_sheet(sheet)
    _title(ws, title, len(headers))
    _header(ws, 3, headers, widths)
    r = 4
    for m in materials:
        rows = m.get(key) or []
        if not rows:
            continue
        start = r
        for row in rows:
            ws.cell(row=r, column=2).value = row.get(row_fields[0], "")
            ws.cell(row=r, column=3).value = _src_str(row.get("source", "")) or _src_str(row.get("required_by", ""))
            ws.cell(row=r, column=4).value = row.get(row_fields[1], "")
            ws.cell(row=r, column=5).value = _fmt(row.get(row_fields[2], ""))
            vc = ws.cell(row=r, column=6)
            vc.value = _canon_verdict(row.get("verdict", ""))
            vc.fill = _fill_for(row.get("verdict", ""))
            ws.cell(row=r, column=7).value = row.get("note", "")
            r += 1
        _merge_label(ws, start, r - 1, _mat_label(m))


def _src_str(v: Any) -> str:
    """Render a provenance value as a readable cell string.

    Review rows may carry `source` either as a plain string or as a provenance
    dict ({source_file, anchor, snippet}); openpyxl cannot write a dict to a
    cell, so collapse the dict into `source_file#anchor — snippet`.
    """
    if v is None:
        return ""
    if isinstance(v, dict):
        head = v.get("source_file") or v.get("file") or ""
        anchor = v.get("anchor") or ""
        snippet = v.get("snippet") or ""
        loc = f"{head}#{anchor}" if head and anchor else (head or anchor)
        return f"{loc} — {snippet}" if loc and snippet else (loc or snippet)
    if isinstance(v, (list, tuple)):
        return "; ".join(_src_str(x) for x in v if x)
    return str(v)


def _fmt(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (dict, list, tuple)):
        return _src_str(v)
    return str(v)


def _merge_label(ws, start: int, end: int, label: str) -> None:
    if end < start:
        return
    if end > start:
        ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
    ws.cell(row=start, column=1).value = label
    ws.cell(row=start, column=1).alignment = Alignment(wrap_text=True, vertical="center")


def _finalize_cells(ws) -> None:
    """Vertically CENTER every populated cell and enable text WRAPPING, keeping
    each cell's existing horizontal alignment and column widths unchanged.

    openpyxl leaves cells with no explicit `vertical` rendering as bottom in
    Excel; data cells were written without alignment, so the whole report body
    was bottom-aligned and long text overflowed. This pass fixes both.
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            cur = cell.alignment
            cell.alignment = Alignment(
                horizontal=cur.horizontal,   # preserve existing left/right alignment
                vertical="center",           # bottom -> middle
                wrap_text=True,              # wrap long text within the column
            )


def build_compliance_report(review_path: Path, out_path: Path) -> Path:
    review = json.loads(Path(review_path).read_text(encoding="utf-8"))
    case_id = review.get("case_id", "")
    materials = review.get("materials") or []

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1) 검토 총괄
    ws = wb.create_sheet("검토 총괄")
    _title(ws, f"성적서 검토 총괄 — Case {case_id}", 8)
    meta = f"PO: {review.get('po_number','—')}   |   MPS: {', '.join(review.get('mps_files') or []) or '—'}"
    if review.get("code_edition_note"):
        meta += f"   |   Code Edition: {review['code_edition_note']}"
    _meta(ws, 2, meta, 8)
    _header(ws, 4,
            ["S/N", "품명", "재질(발주)", "재질(MTC)", "규격(Size)", "Heat No", "수량", "종합 판정"],
            [6, 20, 16, 16, 18, 16, 8, 16])
    r = 5
    for i, m in enumerate(materials, start=1):
        ws.cell(row=r, column=1).value = f"{i:03d}"
        ws.cell(row=r, column=2).value = m.get("item_name", "")
        ws.cell(row=r, column=3).value = m.get("grade_spec", "")
        ws.cell(row=r, column=4).value = m.get("grade_cert", "")
        ws.cell(row=r, column=5).value = m.get("size", "")
        ws.cell(row=r, column=6).value = m.get("heat_no", "")
        ws.cell(row=r, column=7).value = _fmt(m.get("qty", ""))
        vc = ws.cell(row=r, column=8)
        vc.value = _canon_verdict(m.get("verdict", ""))
        vc.fill = _fill_for(m.get("verdict", ""))
        r += 1

    # Phase 1.6 excluded documents (enclosed non-MTC pages) — informational
    # block below the materials table. Missing field on a legacy review.json ->
    # `or []` -> nothing rendered (backward-compatible).
    excluded_docs = review.get("excluded_documents") or []
    if excluded_docs:
        r += 1
        hc = ws.cell(row=r, column=1)
        hc.value = "검토 제외 문서 (완제품 성적서 아님 — 동봉 문서)"
        hc.font = Font(bold=True)
        r += 1
        for d in excluded_docs:
            ws.cell(row=r, column=1).value = "제외됨"
            ws.cell(row=r, column=2).value = d.get("doc_type_ko", d.get("doc_type", ""))
            ws.cell(row=r, column=3).value = d.get("stem", "")
            ws.cell(row=r, column=4).value = d.get("page_range", "")
            ws.cell(row=r, column=5).value = d.get("note", "")
            r += 1

    # 2) 화학성분 검토
    _grouped_sheet(
        wb, "화학성분 검토", "화학성분 검토 (Chemical Composition)",
        ["품명 / Heat", "원소", "규격 출처", "규격 범위", "실측", "판정", "비고"],
        [22, 10, 26, 18, 12, 8, 32],
        materials, "chemistry", ("element", "spec_range", "cert"),
    )
    # 3) 기계적성질 검토
    _grouped_sheet(
        wb, "기계적성질 검토", "기계적 성질 검토 (Mechanical Properties)",
        ["품명 / Heat", "항목", "규격 출처", "규격", "실측", "판정", "비고"],
        [22, 16, 26, 18, 14, 8, 30],
        materials, "mechanical", ("property", "spec", "cert"),
    )
    # 4) 열처리 검토
    _grouped_sheet(
        wb, "열처리 검토", "열처리 검토 (Heat Treatment)",
        ["품명 / Heat", "단계", "규격 출처", "규격 범위", "실측", "판정", "비고"],
        [22, 16, 26, 18, 16, 8, 30],
        materials, "heat_treatment", ("stage", "spec", "cert"),
    )

    # 5) 표기·형식 검토 (doc_checks + nde)
    ws = wb.create_sheet("표기·형식 검토")
    _title(ws, "성적서 표기/형식·NDE 검토 (Document & NDE Check)", 6)
    _header(ws, 3, ["페이지", "검토 위치", "MTC 표기/실측", "마땅한 값/요구", "판정", "비고"],
            [12, 24, 26, 26, 8, 30])
    r = 4
    for m in materials:
        for d in m.get("doc_checks") or []:
            ws.cell(row=r, column=1).value = d.get("page", "")
            ws.cell(row=r, column=2).value = d.get("location", "")
            ws.cell(row=r, column=3).value = _fmt(d.get("mtc_value", ""))
            ws.cell(row=r, column=4).value = _fmt(d.get("expected", ""))
            vc = ws.cell(row=r, column=5)
            vc.value = _canon_verdict(d.get("verdict", ""))
            vc.fill = _fill_for(d.get("verdict", ""))
            ws.cell(row=r, column=6).value = d.get("note", "")
            r += 1
        for n in m.get("nde") or []:
            ws.cell(row=r, column=1).value = ""
            ws.cell(row=r, column=2).value = f"NDE: {n.get('item','')} ({m.get('heat_no','')})"
            ws.cell(row=r, column=3).value = _fmt(n.get("cert", ""))
            ws.cell(row=r, column=4).value = _fmt(n.get("spec", "")) or _src_str(n.get("source", ""))
            vc = ws.cell(row=r, column=5)
            vc.value = _canon_verdict(n.get("verdict", ""))
            vc.fill = _fill_for(n.get("verdict", ""))
            ws.cell(row=r, column=6).value = n.get("note", "")
            r += 1

    # 6) 지적사항 종합
    ws = wb.create_sheet("지적사항 종합")
    _title(ws, "지적사항 종합 (Findings & Recommendations)", 6)
    _header(ws, 3, ["No.", "심각도", "구분", "위치", "내용", "권고 조치"],
            [8, 10, 18, 24, 44, 32])
    r = 4
    for f in review.get("findings") or []:
        ws.cell(row=r, column=1).value = _fmt(f.get("no", ""))
        sc = ws.cell(row=r, column=2)
        sc.value = _canon_severity(f.get("severity", ""))
        sc.fill = _severity_fill(f.get("severity", ""))
        ws.cell(row=r, column=3).value = f.get("category", "")
        ws.cell(row=r, column=4).value = f.get("location", "")
        cc = ws.cell(row=r, column=5)
        cc.value = f.get("content", "")
        cc.alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=6).value = f.get("action", "")
        r += 1

    # Vertically center + wrap every populated cell across all sheets
    # (column widths and horizontal alignment are left unchanged).
    for ws in wb.worksheets:
        _finalize_cells(ws)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


__all__ = ["build_compliance_report"]
