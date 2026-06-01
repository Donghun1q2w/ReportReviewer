"""6-sheet Korean Excel report builder for cert-review-skill.

API
---
    from pathlib import Path
    from scripts.report_builder import build_report

    out_path = build_report(
        case_id="PU2501780",
        findings_path=Path(".cache/_test/PU2501780_findings.json"),
        manifest_path=Path("manifest.json"),
        out_dir=Path("output"),
    )

Constraints
-----------
- C1: no OCR libs (openpyxl only)
- C2/C8: every finding in 지적사항 종합 shows evidence source_file + anchor in 비고
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="FF4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")

_PASS_FILL    = PatternFill("solid", fgColor="FFC6EFCE")
_FAIL_FILL    = PatternFill("solid", fgColor="FFFFC7CE")
_WARN_FILL    = PatternFill("solid", fgColor="FFFFEB9C")
_TITLE_FONT   = Font(bold=True)

# Verdict keyword → fill
_VERDICT_FILL: dict[str, PatternFill] = {
    "PASS":       _PASS_FILL,
    "FAIL":       _FAIL_FILL,
    "WARNING":    _WARN_FILL,
    "조건부 PASS": _WARN_FILL,
}

# Severity → Korean display
_SEVERITY_KO: dict[str, str] = {
    "Reject":        "FAIL",
    "ActionRequired":"FAIL",
    "Question":      "주의",
    "Minor":         "주의",
}

# Verdict severity order for WORST computation (higher = worse)
_VERDICT_RANK: dict[str, int] = {
    "PASS":        0,
    "조건부 PASS": 1,
    "WARNING":     2,
    "FAIL":        3,
}

# Default column widths (matches PU2501780 sample)
_COL_WIDTHS = {
    "A": 6,
    "B": 18,
    "C": 16,
    "D": 22,
    "E": 22,
    "F": 12,
    "G": 10,
    "H": 38,
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_report(
    case_id: str,
    findings_path: Path,
    manifest_path: Path,
    out_dir: Path,
) -> Path:
    """Read <case_id>_findings.json + manifest.json, write
    out_dir/<case_id>_MTC_Review.xlsx, return the path."""

    findings_path = Path(findings_path)
    manifest_path = Path(manifest_path)
    out_dir = Path(out_dir)

    loaded: Any = json.loads(
        findings_path.read_text(encoding="utf-8")
    )
    # Accept either the compare_engine result dict {"findings":[...]} or a bare list.
    if isinstance(loaded, dict):
        findings: list[dict[str, Any]] = loaded.get("findings", []) or []
    else:
        findings = loaded
    manifest: dict[str, Any] = json.loads(
        manifest_path.read_text(encoding="utf-8")
    )

    # Locate case meta from manifest
    case_meta = _find_case(manifest, case_id)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _build_summary(wb, case_id, case_meta, findings)
    _build_chemistry(wb, findings)
    _build_mechanical(wb, findings)
    _build_heat_treatment(wb, findings)
    _build_identification(wb, findings)
    _build_findings_summary(wb, findings)

    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{case_id}_MTC_Review.xlsx"
    wb.save(str(out_path))
    return out_path


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary(
    wb: openpyxl.Workbook,
    case_id: str,
    case_meta: dict[str, Any] | None,
    findings: list[dict[str, Any]],
) -> None:
    """Sheet 1: 검토 총괄"""
    ws = wb.create_sheet("검토 총괄")

    # Derive PO and cert file list from case_meta
    po = case_id
    cert_files: list[str] = []
    if case_meta:
        cert_files = case_meta.get("cert_pdfs", [])

    # A1: title (merged across 8 columns)
    _write_title(ws, f"성적서 검토 총괄 — {case_id} MTC", n_cols=8)

    # A2: meta line (merged)
    cert_list_str = ", ".join(cert_files) if cert_files else "—"
    meta_line = f"발주서: {po}   |   MTC files: {cert_list_str}"
    ws.merge_cells("A2:H2")
    ws["A2"] = meta_line
    ws["A2"].alignment = Alignment(wrap_text=False)

    # A4: header row
    headers = ["S/N", "품명", "재질(발주)", "재질(MTC)", "규격", "Heat No", "수량(발주)", "종합 판정"]
    _write_header_row(ws, row=4, headers=headers)
    _set_col_widths(ws, _COL_WIDTHS)

    # Group findings by (cert_pdf, page_ref, grade, heat_no)
    groups: dict[tuple, list[dict]] = {}
    for f in findings:
        key = (
            f.get("cert_pdf", ""),
            f.get("page_ref", ""),
            f.get("material_grade", ""),
            f.get("heat_no", ""),
        )
        groups.setdefault(key, []).append(f)

    row_idx = 5
    sn = 1
    for (cert_pdf, page_ref, grade, heat_no), group_findings in groups.items():
        worst = _worst_verdict(group_findings)
        fail_summaries = [
            f["issue_summary"]
            for f in group_findings
            if _finding_verdict(f) == "FAIL"
        ]
        verdict_str = worst
        if fail_summaries:
            verdict_str = f"{worst} ({'; '.join(fail_summaries[:2])})"

        # item_name from findings detail or page_ref
        item_name = _extract_item_name(group_findings)
        grade_ordered = grade or "—"
        heat_str = heat_no or "—"

        ws.cell(row=row_idx, column=1).value = f"{sn:03d}"
        ws.cell(row=row_idx, column=2).value = item_name
        ws.cell(row=row_idx, column=3).value = grade_ordered  # grade as ordered (from MPS)
        ws.cell(row=row_idx, column=4).value = grade_ordered  # grade as in MTC
        ws.cell(row=row_idx, column=5).value = page_ref or "—"
        ws.cell(row=row_idx, column=6).value = heat_str
        ws.cell(row=row_idx, column=7).value = "—"
        verdict_cell = ws.cell(row=row_idx, column=8)
        verdict_cell.value = verdict_str
        verdict_cell.fill = _verdict_fill(worst)

        sn += 1
        row_idx += 1

    # Footer note row (merged, italic)
    if row_idx > 5:
        ws.merge_cells(f"A{row_idx}:H{row_idx}")
        ws.cell(row=row_idx, column=1).value = (
            "※ 상기 판정은 성적서 검토 결과이며, 표기 오류는 발행처 정정본 재발급이 필요합니다."
        )
        ws.cell(row=row_idx, column=1).font = Font(italic=True)


def _build_chemistry(
    wb: openpyxl.Workbook,
    findings: list[dict[str, Any]],
) -> None:
    """Sheet 2: 화학성분 검토"""
    ws = wb.create_sheet("화학성분 검토")

    _write_title(ws, "화학성분 검토 (Chemical Composition)", n_cols=7)
    headers = ["품명 / Heat", "원소", "규격 출처", "규격 범위(%)", "실측(%)", "판정", "비고"]
    _write_header_row(ws, row=3, headers=headers)
    _set_col_widths(ws, _COL_WIDTHS)

    chem_findings = [f for f in findings if f.get("category") == "Chemistry"]

    row_idx = 4
    # Group by (item_name, heat_no) for merged A column
    groups: dict[tuple, list[dict]] = {}
    for f in chem_findings:
        key = (_extract_item_name([f]), f.get("heat_no", ""))
        groups.setdefault(key, []).append(f)

    for (item_name, heat_no), group in groups.items():
        group_start = row_idx
        for f in group:
            structured = f.get("structured", {}) or {}
            element    = structured.get("element", "")
            spec_src   = _evidence_ref(f)
            spec_range = _format_range(structured)
            measured   = structured.get("measured", "")

            verdict = _finding_verdict(f)

            # A column: only written on first row of group, merged later
            ws.cell(row=row_idx, column=2).value = element
            ws.cell(row=row_idx, column=3).value = spec_src
            ws.cell(row=row_idx, column=4).value = spec_range
            ws.cell(row=row_idx, column=5).value = str(measured) if measured != "" else ""
            v_cell = ws.cell(row=row_idx, column=6)
            v_cell.value = verdict
            v_cell.fill = _verdict_fill(verdict)
            ws.cell(row=row_idx, column=7).value = f.get("issue_summary", "")

            row_idx += 1

        group_end = row_idx - 1
        label = f"{item_name}\n({heat_no})" if heat_no else item_name
        if group_start == group_end:
            ws.cell(row=group_start, column=1).value = label
            ws.cell(row=group_start, column=1).alignment = Alignment(
                wrap_text=True, vertical="center"
            )
        else:
            ws.merge_cells(
                start_row=group_start, start_column=1,
                end_row=group_end, end_column=1,
            )
            ws.cell(row=group_start, column=1).value = label
            ws.cell(row=group_start, column=1).alignment = Alignment(
                wrap_text=True, vertical="center"
            )


def _build_mechanical(
    wb: openpyxl.Workbook,
    findings: list[dict[str, Any]],
) -> None:
    """Sheet 3: 기계적성질 검토"""
    ws = wb.create_sheet("기계적성질 검토")

    _write_title(ws, "기계적 성질 검토 (Mechanical Properties)", n_cols=7)
    headers = ["품명 / Heat", "항목", "규격 출처", "규격", "실측", "판정", "비고"]
    _write_header_row(ws, row=3, headers=headers)
    _set_col_widths(ws, _COL_WIDTHS)

    mech_findings = [f for f in findings if f.get("category") == "Mechanical"]

    row_idx = 4
    groups: dict[tuple, list[dict]] = {}
    for f in mech_findings:
        key = (_extract_item_name([f]), f.get("heat_no", ""))
        groups.setdefault(key, []).append(f)

    for (item_name, heat_no), group in groups.items():
        group_start = row_idx
        for f in group:
            structured = f.get("structured", {}) or {}
            item_label = structured.get("property", "") or f.get("issue_summary", "")
            spec_src   = _evidence_ref(f)
            spec_str   = structured.get("spec", "")
            measured   = structured.get("measured", "")

            verdict = _finding_verdict(f)

            ws.cell(row=row_idx, column=2).value = item_label
            ws.cell(row=row_idx, column=3).value = spec_src
            ws.cell(row=row_idx, column=4).value = str(spec_str)
            ws.cell(row=row_idx, column=5).value = str(measured) if measured != "" else ""
            v_cell = ws.cell(row=row_idx, column=6)
            v_cell.value = verdict
            v_cell.fill = _verdict_fill(verdict)
            ws.cell(row=row_idx, column=7).value = f.get("issue_summary", "")

            row_idx += 1

        group_end = row_idx - 1
        label = f"{item_name}\n({heat_no})" if heat_no else item_name
        if group_start == group_end:
            ws.cell(row=group_start, column=1).value = label
            ws.cell(row=group_start, column=1).alignment = Alignment(
                wrap_text=True, vertical="center"
            )
        else:
            ws.merge_cells(
                start_row=group_start, start_column=1,
                end_row=group_end, end_column=1,
            )
            ws.cell(row=group_start, column=1).value = label
            ws.cell(row=group_start, column=1).alignment = Alignment(
                wrap_text=True, vertical="center"
            )


def _build_heat_treatment(
    wb: openpyxl.Workbook,
    findings: list[dict[str, Any]],
) -> None:
    """Sheet 4: 열처리 검토"""
    ws = wb.create_sheet("열처리 검토")

    _write_title(ws, "열처리 검토 (Heat Treatment)", n_cols=7)
    headers = ["품명 / Heat", "단계", "규격 출처", "규격 범위", "실측", "판정", "비고"]
    _write_header_row(ws, row=3, headers=headers)
    _set_col_widths(ws, _COL_WIDTHS)

    ht_findings = [f for f in findings if f.get("category") == "HeatTreatment"]

    row_idx = 4
    groups: dict[tuple, list[dict]] = {}
    for f in ht_findings:
        key = (_extract_item_name([f]), f.get("heat_no", ""))
        groups.setdefault(key, []).append(f)

    for (item_name, heat_no), group in groups.items():
        group_start = row_idx
        for f in group:
            structured = f.get("structured", {}) or {}
            stage    = structured.get("stage", "") or f.get("issue_summary", "")
            spec_src = _evidence_ref(f)
            spec_rng = structured.get("spec", "")
            measured = structured.get("measured", "")
            note     = f.get("issue_summary", "")

            verdict = _finding_verdict(f)

            ws.cell(row=row_idx, column=2).value = stage
            ws.cell(row=row_idx, column=3).value = spec_src
            ws.cell(row=row_idx, column=4).value = str(spec_rng)
            ws.cell(row=row_idx, column=5).value = str(measured) if measured != "" else ""
            v_cell = ws.cell(row=row_idx, column=6)
            v_cell.value = verdict
            v_cell.fill = _verdict_fill(verdict)
            ws.cell(row=row_idx, column=7).value = str(note)

            row_idx += 1

        group_end = row_idx - 1
        label = f"{item_name}\n({heat_no})" if heat_no else item_name
        if group_start == group_end:
            ws.cell(row=group_start, column=1).value = label
            ws.cell(row=group_start, column=1).alignment = Alignment(
                wrap_text=True, vertical="center"
            )
        else:
            ws.merge_cells(
                start_row=group_start, start_column=1,
                end_row=group_end, end_column=1,
            )
            ws.cell(row=group_start, column=1).value = label
            ws.cell(row=group_start, column=1).alignment = Alignment(
                wrap_text=True, vertical="center"
            )


def _build_identification(
    wb: openpyxl.Workbook,
    findings: list[dict[str, Any]],
) -> None:
    """Sheet 5: 표기·형식 검토"""
    ws = wb.create_sheet("표기·형식 검토")

    _write_title(ws, "성적서 표기/형식 검토 (Document Form Check)", n_cols=6)
    headers = ["페이지", "검토 위치", "MTC 표기", "마땅한 값", "판정", "비고"]
    _write_header_row(ws, row=3, headers=headers)

    # Column widths for 6-col sheet
    col_widths_6 = {"A": 14, "B": 22, "C": 26, "D": 26, "E": 10, "F": 38}
    _set_col_widths(ws, col_widths_6)

    id_findings = [
        f for f in findings
        if f.get("category") in ("Identification", "DocumentError")
    ]

    row_idx = 4
    for f in id_findings:
        structured  = f.get("structured", {}) or {}
        page_ref    = f.get("page_ref", "")
        location    = structured.get("location", "") or f.get("issue_summary", "")
        mtc_value   = structured.get("mtc_value", "") or ""
        correct_val = structured.get("correct_value", "") or ""
        note        = f.get("issue_summary", "")

        verdict = _finding_verdict(f)

        ws.cell(row=row_idx, column=1).value = page_ref
        ws.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=row_idx, column=2).value = location
        ws.cell(row=row_idx, column=3).value = mtc_value
        ws.cell(row=row_idx, column=4).value = correct_val
        v_cell = ws.cell(row=row_idx, column=5)
        v_cell.value = verdict
        v_cell.fill = _verdict_fill(verdict)
        ws.cell(row=row_idx, column=6).value = note

        row_idx += 1


def _build_findings_summary(
    wb: openpyxl.Workbook,
    findings: list[dict[str, Any]],
) -> None:
    """Sheet 6: 지적사항 종합 (ALL findings, C2/C8: evidence source in 출처 column)."""
    ws = wb.create_sheet("지적사항 종합")

    _write_title(ws, "지적사항 종합 (Findings & Recommendations)", n_cols=8)
    headers = [
        "No.", "심각도", "구분", "위치", "내용", "권고 조치",
        "Code Edition", "출처",
    ]
    _write_header_row(ws, row=3, headers=headers)

    # Wider columns for this sheet
    col_widths_8 = {
        "A": 6, "B": 10, "C": 18, "D": 22,
        "E": 40, "F": 30, "G": 20, "H": 38,
    }
    _set_col_widths(ws, col_widths_8)

    row_idx = 4
    for i, f in enumerate(findings, start=1):
        severity_raw = f.get("severity", "Minor")
        severity_ko  = _SEVERITY_KO.get(severity_raw, "주의")
        category     = f.get("category", "Other")
        page_ref     = f.get("page_ref", "")
        summary      = f.get("issue_summary", "")
        details_raw  = f.get("details", "")
        details_str  = (
            details_raw if isinstance(details_raw, str)
            else json.dumps(details_raw, ensure_ascii=False)
        )
        content      = details_str if details_str else summary
        action       = f.get("required_action", "—")

        # Code edition mismatch note
        code_edition = _code_edition_note(f)

        # C2/C8: evidence provenance — source_file + anchor
        source_str = _evidence_provenance(f)

        # verdict for colour
        verdict = _finding_verdict(f)

        ws.cell(row=row_idx, column=1).value = i
        sev_cell = ws.cell(row=row_idx, column=2)
        sev_cell.value = severity_ko
        sev_cell.fill = _verdict_fill(verdict)
        ws.cell(row=row_idx, column=3).value = category
        ws.cell(row=row_idx, column=4).value = page_ref
        content_cell = ws.cell(row=row_idx, column=5)
        content_cell.value = content
        content_cell.alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=6).value = action
        ws.cell(row=row_idx, column=7).value = code_edition
        ws.cell(row=row_idx, column=8).value = source_str  # C2/C8 provenance

        row_idx += 1


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_title(ws: Any, title: str, n_cols: int) -> None:
    """Write a bold merged title in row 1."""
    last_col = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last_col}1")
    cell = ws["A1"]
    cell.value = title
    cell.font = _TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20


def _write_header_row(ws: Any, row: int, headers: list[str]) -> None:
    """Write a styled header row."""
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = h
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22


def _set_col_widths(ws: Any, widths: dict[str, float]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _finding_verdict(f: dict[str, Any]) -> str:
    """Map finding severity → display verdict string."""
    sev = f.get("severity", "Minor")
    if sev in ("Reject", "ActionRequired"):
        return "FAIL"
    if sev == "Question":
        return "WARNING"
    # Minor — check if explicitly marked pass
    if "PASS" in str(f.get("issue_summary", "")).upper():
        return "PASS"
    return "PASS"  # Minor → treat as PASS note


def _worst_verdict(findings: list[dict[str, Any]]) -> str:
    verdicts = [_finding_verdict(f) for f in findings]
    return max(verdicts, key=lambda v: _VERDICT_RANK.get(v, 0), default="PASS")


def _verdict_fill(verdict: str) -> PatternFill:
    return _VERDICT_FILL.get(verdict, PatternFill())


def _evidence_ref(f: dict[str, Any]) -> str:
    """Return '<source_file>:<anchor>' from first ref_code evidence, else first evidence."""
    evidence: list[dict] = f.get("evidence", [])
    # prefer ref_code channel
    for ev in evidence:
        if ev.get("channel") == "ref_code":
            sf = ev.get("source_file", "")
            anchor = ev.get("anchor", "")
            return f"{sf}:{anchor}" if anchor else sf
    # fallback: first evidence
    if evidence:
        ev = evidence[0]
        sf = ev.get("source_file", "")
        anchor = ev.get("anchor", "")
        return f"{sf}:{anchor}" if anchor else sf
    return ""


def _evidence_provenance(f: dict[str, Any]) -> str:
    """C2/C8: Return 'source_file:anchor' from first evidence item."""
    evidence: list[dict] = f.get("evidence", [])
    if not evidence:
        return "—"
    ev = evidence[0]
    sf = ev.get("source_file", "")
    anchor = ev.get("anchor", "")
    return f"{sf}:{anchor}" if anchor else sf


def _code_edition_note(f: dict[str, Any]) -> str:
    """Return code edition mismatch note if present in evidence."""
    evidence: list[dict] = f.get("evidence", [])
    editions: list[str] = []
    for ev in evidence:
        ch = ev.get("channel", "")
        src = ev.get("source_file", "")
        if ch == "ref_code" and src:
            editions.append(f"ref_code: {src}")
        elif ch == "mps" and src:
            editions.append(f"MPS: {src}")
    if len(editions) >= 2:
        return " / ".join(editions)
    return editions[0] if editions else "—"


def _format_range(details: dict[str, Any]) -> str:
    """Format spec range as 'min–max', 'max X', or 'min X'."""
    mn = details.get("min")
    mx = details.get("max")
    if mn is not None and mx is not None:
        return f"{mn}–{mx}"
    if mx is not None:
        return f"max {mx}"
    if mn is not None:
        return f"min {mn}"
    return details.get("spec", "")


def _extract_item_name(findings: list[dict[str, Any]]) -> str:
    """Best-effort item name from first finding structured fields or page_ref."""
    for f in findings:
        structured = f.get("structured", {}) or {}
        name = structured.get("item_name") or structured.get("item") or ""
        if name:
            return str(name)
        # Try page_ref prefix
        pr = f.get("page_ref", "")
        if pr:
            return pr
    return "—"


def _find_case(
    manifest: dict[str, Any], case_id: str
) -> dict[str, Any] | None:
    for case in manifest.get("cases", []):
        if case.get("case_id") == case_id:
            return case
    return None
